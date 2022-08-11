import pandas as pd
from openpyxl import load_workbook

wbname = "large10pk.xlsx"
wb = load_workbook(wbname)
ws = wb.active

print("start")

data =[]
for col in ws['B']:
	data.append(col.value)
setdata = set(data)

print('breaking df')
df1 = pd.read_csv('chunk' + str(batch_no)+'.csv', usecols = ["address"])
setdf1 = set(df1)

print (data)
print("read")