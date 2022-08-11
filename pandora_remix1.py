import pandas as pd
from openpyxl import load_workbook, Workbook


for h in range(1,13):
	wbname = "pk" + str(h )+".xlsx"
	wb = load_workbook(wbname)
	ws = wb.active
	savewb = Workbook()
	savews = savewb.active
	

	print("start")
	print(wbname)

	data =[]
	for col in ws['B']:
		data.append(col.value)
	setdata = set(data)

	print('breaking df')
	
	for i in range(1,5):
		df1 = pd.read_csv('chunk' +str(i) + '.csv' , usecols = ["address"])
		setdf1 = set(df1["address"])
		match = setdf1.intersection(setdata)
		print(match)
		if len(match) > 0:
			listmatch = list (match)
			row = [listmatch, wbname]
			savews.append(row)
			savewb.save("balancer.xlsx")
		print('next batch is ' + str(i))
			
print ("read")
