import pandas as pd
from openpyxl import load_workbook

wbname = "large10pk.xlsx"
wb = load_workbook(wbname)
ws = wb.active

print("start")

data =[]
for col in ws['B']:
	data.append(col.value)
setdata = set(data)

print('breaking df')

chunk_size = 1000000
batch_no = 1
for chunk in pd.read_csv ("Ethbal00.csv", usecols =["address"], chunksize = chunk_size):
	chunk.to_csv('chunk' + str(batch_no)+'.csv', index='false')
	df1 = pd.read_csv('chunk' + str(batch_no)+'.csv', usecols = ["address"])
	setdf1 = set(df1["address"])
	match = setdf1.intersection(setdata)
	print(match)
	batch_no += 1
	print('next batch is ' + str(batch_no))
print("read")